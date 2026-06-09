#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX_5 = Path(r"g:\Outros computadores\Meu laptop\G8 Representações\PANTANEIRO\Tabela de Preços Completa (281400) -  Região Sulsudeste - Comissão 5%.xlsx")
XLSX_7 = Path(r"g:\Outros computadores\Meu laptop\G8 Representações\PANTANEIRO\Tabela de Preços Completa (301400) -  Região Sulsudeste - Comissão 7%.xlsx")

FILES_5 = [
    ROOT / "public" / "pantaneiro5.html",
    ROOT / "public" / "b2b-pantaneiro5.html",
]
FILES_7 = [
    ROOT / "public" / "pantaneiro7.html",
    ROOT / "public" / "b2b-pantaneiro7.html",
]


def resolve_cell_value(ws, value):
    if isinstance(value, str) and value.startswith("="):
        ref = value.lstrip("=").replace("$", "")
        target = ws[ref].value if ref else None
        return resolve_cell_value(ws, target) if isinstance(target, str) and target.startswith("=") else target
    return value


def parse_tamanhos(tam_str: str) -> list[str]:
    if not tam_str:
        return ["ÚNICO"]
    s = normalize_text(tam_str)
    s = re.sub(r"(?i)^.{0,2}nico", "ÚNICO", s, count=1)
    if s.upper() in {"ÚNICO", "UNICO"}:
        return ["ÚNICO"]
    if "BOTAS" in s.upper() or "CILINDRADAS" in s.upper() or s.upper().startswith("AT"):
        return [s]
    if re.search(r"\d+/\d+", s):
        return [p.strip() for p in re.split(r"-", s) if p.strip()]
    if " - " in s:
        return [p.strip() for p in s.split(" - ") if p.strip()]
    if "/" in s:
        return [p.strip() for p in s.split("/") if p.strip()]
    return [s]


def normalize_text(value) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    s = re.sub(r"\s+", " ", s)
    s = s.replace("\ufffd", "")
    replacements = {
        "\xda": "Ú",
        "\xfa": "ú",
        "\xc7": "Ç",
        "\xe7": "ç",
        "\xc1": "Á",
        "\xe1": "á",
        "\xc9": "É",
        "\xe9": "é",
        "\xcd": "Í",
        "\xed": "í",
        "\xd3": "Ó",
        "\xf3": "ó",
        "\xc3": "Ã",
        "\xe3": "ã",
        "\xd5": "Õ",
        "\xf5": "õ",
        "\xc2": "Â",
        "\xe2": "â",
        "\xca": "Ê",
        "\xea": "ê",
        "\xd4": "Ô",
        "\xf4": "ô",
    }
    for src, dst in replacements.items():
        s = s.replace(src, dst)
    return s


def format_ref(ref) -> str:
    if isinstance(ref, (int, float)):
        f = float(ref)
        if f == int(f):
            return str(int(f))
        return str(ref).strip()
    return str(ref).strip()


def load_products(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    wb_formulas = openpyxl.load_workbook(path, data_only=False)
    ws = wb.active
    ws_formulas = wb_formulas.active
    products = []
    category = None

    for row in range(12, ws.max_row + 1):
        ref = ws.cell(row, 1).value
        desc = ws.cell(row, 3).value
        tam = ws.cell(row, 4).value
        preco = ws.cell(row, 5).value

        if ref and not desc and not tam:
            category = normalize_text(ref).title()
            if "Acess" in category:
                category = "Urbano - Acessórios"
            continue

        if not ref or not desc:
            continue

        if preco is None:
            continue

        tam_value = resolve_cell_value(ws_formulas, ws_formulas.cell(row, 4).value)
        if tam_value is None:
            tam_value = tam

        products.append(
            {
                "CATEGORIA": category,
                "REFERENCIA": format_ref(ref),
                "DESCRIÇÃO": normalize_text(desc).upper(),
                "TAMANHOS": parse_tamanhos(normalize_text(tam_value)),
                "PRECO": round(float(preco), 2),
            }
        )

    return products


def to_js_block(products: list[dict]) -> str:
    lines = ["// ========== INÍCIO DA LISTA DE PRODUTOS ATUALIZADA ==========", "window.produtosData = ["]
    for idx, product in enumerate(products):
        lines.append("  {")
        lines.append(f'    CATEGORIA: {json.dumps(product["CATEGORIA"], ensure_ascii=False)},')
        lines.append(f'    REFERENCIA: {json.dumps(product["REFERENCIA"], ensure_ascii=False)},')
        lines.append(f'    DESCRIÇÃO: {json.dumps(product["DESCRIÇÃO"], ensure_ascii=False)},')
        if len(product["TAMANHOS"]) == 1:
            lines.append(f'    TAMANHOS: {json.dumps(product["TAMANHOS"], ensure_ascii=False)},')
        else:
            lines.append("    TAMANHOS: [")
            for size in product["TAMANHOS"]:
                lines.append(f'      {json.dumps(size, ensure_ascii=False)},')
            lines.append("    ],")
        lines.append(f'    PRECO: {product["PRECO"]},')
        lines.append("  }," if idx < len(products) - 1 else "  }")
    lines.append("];")
    return "\n".join(lines)


def replace_products_in_html(html_path: Path, js_block: str) -> None:
    content = html_path.read_text(encoding="utf-8")
    pattern = r"// ========== INÍCIO DA LISTA DE PRODUTOS ATUALIZADA ==========\nwindow\.produtosData = \[.*?\];"
    if not re.search(pattern, content, flags=re.DOTALL):
        raise RuntimeError(f"Bloco produtosData não encontrado em {html_path}")
    updated = re.sub(pattern, js_block, content, count=1, flags=re.DOTALL)
    html_path.write_text(updated, encoding="utf-8")
    print(f"Atualizado: {html_path}")


def main():
    products_5 = load_products(XLSX_5)
    products_7 = load_products(XLSX_7)
    js_5 = to_js_block(products_5)
    js_7 = to_js_block(products_7)

    (ROOT / "scripts" / "produtos-pantaneiro5.js").write_text(js_5 + "\n", encoding="utf-8")
    (ROOT / "scripts" / "produtos-pantaneiro7.js").write_text(js_7 + "\n", encoding="utf-8")

    for file_path in FILES_5:
        replace_products_in_html(file_path, js_5)
    for file_path in FILES_7:
        replace_products_in_html(file_path, js_7)

    print(f"Pantaneiro 5: {len(products_5)} produtos")
    print(f"Pantaneiro 7: {len(products_7)} produtos")


if __name__ == "__main__":
    main()
